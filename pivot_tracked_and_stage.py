#!/usr/bin/env python3
"""
pivot_tracked_and_stage.py

SCRIPT ORCHESTRATEUR COMPLET

Ce script regroupe désormais **3 grandes étapes**, sans modifier leur logique interne :

1️⃣ Téléchargement des Tracked Entity Instances (ex-download_tracked.py)
2️⃣ Téléchargement des Events / Program Stages (ex-download.py)
3️⃣ Pivot + génération du fichier Excel (code existant)

⚠️ IMPORTANT :
- Aucune logique n’a été modifiée
- Les fonctions ont simplement été regroupées
- Les paramètres restent configurés via CLI > .env > défauts
"""

# ==========================================================
# ======================= IMPORTS ==========================
# ==========================================================

import argparse
import json
import math
import os
import sys
from pathlib import Path
from urllib.parse import urlencode

import pandas as pd
import requests
from dotenv import load_dotenv
from tqdm import tqdm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ==========================================================
# ================== FONCTIONS COMMUNES ===================
# ==========================================================

def build_url(base_url, params):
    """
    Construit une URL complète DHIS2 à partir :
    - de l’URL de base
    - d’un dictionnaire de paramètres
    """
    query_string = urlencode(params)
    return f"{base_url}?{query_string}" if query_string else base_url


# ==========================================================
# ========== 1️⃣ DOWNLOAD TRACKED ENTITY INSTANCES =========
# ==========================================================

def download_dhis2_tracked(base_url, params, output_file, token):
    """
    Télécharge le CSV des Tracked Entity Instances depuis DHIS2
    (code IDENTIQUE à download_tracked.py)
    """
    full_url = build_url(base_url, params)
    print(f"URL générée (TRACKED) : {full_url}")

    headers = {"Authorization": f"ApiToken {token}"}

    try:
        response = requests.get(full_url, headers=headers, stream=True, timeout=60)
        response.raise_for_status()

        total_size = int(response.headers.get("content-length", 0))

        with open(output_file, "wb") as f, tqdm(
            total=total_size,
            unit="B",
            unit_scale=True,
            unit_divisor=1024,
            desc="Téléchargement Tracked"
        ) as bar:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    bar.update(len(chunk))

        print(f"✅ Tracked Entity Instances téléchargés : {output_file}")

    except requests.exceptions.RequestException as e:
        print(f"❌ Erreur téléchargement tracked : {e}")
        sys.exit(1)


# ==========================================================
# ========== 2️⃣ DOWNLOAD EVENTS / PROGRAM STAGES ==========
# ==========================================================

def download_dhis2_events(base_url, params, output_file, token):
    """
    Télécharge le CSV des Events (Program Stages)
    (code IDENTIQUE à download.py)
    """
    full_url = build_url(base_url, params)
    print(f"URL générée (EVENTS) : {full_url}")

    headers = {"Authorization": f"ApiToken {token}"}

    try:
        response = requests.get(full_url, headers=headers, stream=True, timeout=60)
        response.raise_for_status()

        total_size = int(response.headers.get("content-length", 0))

        with open(output_file, "wb") as f, tqdm(
            total=total_size,
            unit="B",
            unit_scale=True,
            unit_divisor=1024,
            desc="Téléchargement Events"
        ) as bar:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
                    bar.update(len(chunk))

        print(f"✅ Events téléchargés : {output_file}")

    except requests.exceptions.RequestException as e:
        print(f"❌ Erreur téléchargement events : {e}")
        sys.exit(1)


# ==========================================================
# ======================= ÉTAT =============================
# ==========================================================

def load_state(state_file: str) -> dict:
    if os.path.exists(state_file):
        with open(state_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"completed_stages": []}

def save_state(state_file: str, state: dict):
    Path(state_file).parent.mkdir(parents=True, exist_ok=True)
    with open(state_file, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def clear_state(state_file: str):
    if os.path.exists(state_file):
        os.remove(state_file)

# ==========================================================
# ======================= API DHIS2 ========================
# ==========================================================

def safe_get(url: str, token: str, params: dict | None = None) -> dict:
    headers = {"Authorization": f"ApiToken {token}"}
    response = requests.get(url, headers=headers, params=params, timeout=60)
    response.raise_for_status()
    return response.json()

def get_program_stages(base_url: str, token: str, program_id: str) -> list[dict]:
    url = f"{base_url}/programs/{program_id}/programStages.json"
    params = {"fields": "id,displayName,sortOrder", "paging": "false"}
    data = safe_get(url, token, params)
    return sorted(data.get("programStages", []), key=lambda s: s.get("sortOrder", math.inf))

def get_stage_dataelements(base_url: str, token: str, stage_id: str) -> list[str]:
    url = f"{base_url}/programStages/{stage_id}.json"
    params = {
        "fields": "programStageDataElements[dataElement[id,displayName],sortOrder]",
        "paging": "false"
    }
    data = safe_get(url, token, params)
    elements = sorted(
        data.get("programStageDataElements", []),
        key=lambda x: x.get("sortOrder", math.inf)
    )
    return [e["dataElement"]["displayName"] for e in elements if e.get("dataElement")]

# ==========================================================
# ===================== MAPPING =============================
# ==========================================================

def load_mapping(mapping_file: str) -> dict:
    if os.path.exists(mapping_file):
        with open(mapping_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_mapping(mapping_file: str, mapping: dict):
    Path(mapping_file).parent.mkdir(parents=True, exist_ok=True)
    with open(mapping_file, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

def build_de_mapping_from_api(base_url: str, token: str) -> dict:
    url = f"{base_url}/dataElements.json"
    params = {"paging": "false", "fields": "id,displayName"}
    data = safe_get(url, token, params)
    return {de["id"]: de["displayName"] for de in data.get("dataElements", [])}

# ==========================================================
# ======================= PIVOTS ============================
# ==========================================================

def pivot_tracked_df(input_file: str, aggfunc: str) -> pd.DataFrame:
    df = pd.read_csv(input_file, dtype=str)
    pivot = df.pivot_table(
        index="trackedEntityInstance",
        columns="displayName",
        values="value",
        aggfunc=aggfunc
    ).reset_index().fillna("")
    pivot.insert(
        pivot.columns.get_loc("trackedEntityInstance") + 1,
        "ID",
        range(1, len(pivot) + 1)
    )
    return pivot

def run_pivot_and_excel(
    tracked_input: str,
    stage_input: str,
    output: str,
    base_url: str,
    token: str,
    program: str,
    aggfunc: str,
    mapping_file: str,
    state_file: str,
    strict: bool,
):
    """
    Exécute TOUTE la logique existante de pivot et génération Excel.
    ⚠️ Code IDENTIQUE à l’existant, simplement encapsulé.
    """

    # ======================
    # Chargement de l’état
    # ======================
    state = load_state(state_file)

    # ======================
    # Pivot Tracked Entities
    # ======================
    tracked_df = pivot_tracked_df(tracked_input, aggfunc)
    print(f"▶️ Pivot du premier onglet : {len(tracked_df)} lignes")

    cols = list(tracked_df.columns)
    serial_cols = [c for c in cols if "_serial_number" in str(c).lower()]
    date_cols = [c for c in cols if "_date_" in str(c).lower()]
    parent_cols = [c for c in cols if "_parent_consent" in str(c).lower()]

    if strict:
        ordered_cols = ["trackedEntityInstance"] + serial_cols + ["ID"] + date_cols + parent_cols
    else:
        ordered_cols = ["trackedEntityInstance"] + serial_cols + ["ID"]
        remaining = [c for c in cols if c not in ordered_cols]
        ordered_cols += sorted(remaining)

    tracked_df = tracked_df[ordered_cols]

    # ======================
    # Écriture premier onglet
    # ======================
    if os.path.exists(output):
        wb = load_workbook(output)
        if "TrackedEntities" not in wb.sheetnames:
            writer = pd.ExcelWriter(output, engine="openpyxl", mode="a", if_sheet_exists="overlay")
            write_with_progress(tracked_df, writer, "TrackedEntities")
            writer.close()
        else:
            print("⏩ Onglet TrackedEntities déjà présent")
    else:
        writer = pd.ExcelWriter(output, engine="openpyxl")
        write_with_progress(tracked_df, writer, "TrackedEntities")
        writer.close()

    # ======================
    # Réouverture pour stages
    # ======================
    writer = pd.ExcelWriter(output, engine="openpyxl", mode="a", if_sheet_exists="overlay")

    df = pd.read_csv(stage_input, dtype=str)

    # ======================
    # Mapping dataElements
    # ======================
    mapping = load_mapping(mapping_file)
    missing = [u for u in df["dataElement"].unique() if u not in mapping]

    if missing:
        api_map = build_de_mapping_from_api(base_url, token)
        for uid in missing:
            mapping[uid] = api_map.get(uid, uid)
        save_mapping(mapping_file, mapping)

    df["dataElement"] = df["dataElement"].map(mapping)

    # ======================
    # Pivot global events
    # ======================
    pivot = df.pivot_table(
        index="enrollment",
        columns="dataElement",
        values="value",
        aggfunc=aggfunc
    ).reset_index().fillna("")

    pivot.insert(
        pivot.columns.get_loc("enrollment") + 1,
        "ID",
        range(1, len(pivot) + 1)
    )

    stages = get_program_stages(base_url, token, program)

    try:
        for stage in stages:
            stage_name = stage["displayName"][:31]

            if stage_name in state["completed_stages"]:
                print(f"⏩ Déjà traité : {stage_name}")
                continue

            print(f"▶️ Traitement : {stage_name}")

            elements = get_stage_dataelements(base_url, token, stage["id"])
            cols = ["enrollment", "ID"] + [c for c in elements if c in pivot.columns]
            sheet_df = pivot[cols]

            if sheet_df.columns.tolist() == ["enrollment", "ID"]:
                print("⚠️ Stage vide ignoré")
                state["completed_stages"].append(stage_name)
                save_state(state_file, state)
                continue

            sheet_df.to_excel(writer, sheet_name=stage_name, index=False)
            auto_adjust_column_width(writer.sheets[stage_name])

            state["completed_stages"].append(stage_name)
            save_state(state_file, state)

            print(f"✅ Terminé : {stage_name}")

    finally:
        writer.close()

    clear_state(state_file)
    print("\n🎉 Fichier Excel généré avec succès")

# ==========================================================
# ===================== EXCEL UTIL =========================
# ==========================================================

def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            value = str(cell.value) if cell.value else ""
            max_length = max(max_length, len(value))
        ws.column_dimensions[column].width = max_length + 2

def write_with_progress(df: pd.DataFrame, writer, sheet_name: str, chunk_size: int = 50):
    total = len(df)
    df.head(0).to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    for start in range(0, total, chunk_size):
        end = min(start + chunk_size, total)
        chunk = df.iloc[start:end]
        for r_idx, row in enumerate(chunk.values, start=start + 2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        progress = (end / total) * 100
        sys.stdout.write(f"\r   → Lignes {start+1} à {end} / {total} ({progress:.1f}%)")
        sys.stdout.flush()

    auto_adjust_column_width(ws)
    print(f"\n✅ Onglet '{sheet_name}' écrit avec succès\n")

# ==========================================================
# ========================= MAIN ============================
# ==========================================================

def main():
    load_dotenv()

    parser = argparse.ArgumentParser(description="Pipeline DHIS2 complet")
    parser.add_argument("--skip-download", action="store_true")
    parser.add_argument("--only-download", action="store_true")
    parser.add_argument("--only-pivot", action="store_true")

    args, _ = parser.parse_known_args()

    # Sécurité : flags incompatibles
    if args.only_download and args.only_pivot:
        parser.error("--only-download et --only-pivot sont incompatibles")

    # ======================================================
    # DÉTERMINATION DU MODE D’EXÉCUTION
    # ======================================================

    do_download = True
    do_pivot = True

    if args.only_download:
        do_pivot = False

    if args.only_pivot:
        do_download = False

    if args.skip_download:
        do_download = False

    # ======================================================
    # ÉTAPE 1 : DOWNLOAD
    # ======================================================
    if do_download:
        print("\n📥 ÉTAPE : Téléchargement DHIS2\n")

        download_dhis2_tracked(
            base_url=os.getenv("TRACKED_BASE_URL"),
            params={
                "program": os.getenv("TRACKED_PROGRAM"),
                "programStartDate": os.getenv("TRACKED_PROGRAM_START_DATE"),
                "programEndDate": os.getenv("TRACKED_PROGRAM_END_DATE"),
                "ouMode": os.getenv("TRACKED_OU_MODE"),
                "format": os.getenv("TRACKED_FORMAT"),
            },
            output_file=os.getenv("TRACKED_OUTPUT"),
            token=os.getenv("PIVOT_TOKEN"),
        )

        download_dhis2_events(
            base_url=os.getenv("DOWNLOAD_BASE_URL"),
            params={
                "orgUnit": os.getenv("DOWNLOAD_ORG_UNIT"),
                "program": os.getenv("DOWNLOAD_PROGRAM"),
                "startDate": os.getenv("DOWNLOAD_START_DATE"),
                "endDate": os.getenv("DOWNLOAD_END_DATE"),
                "ouMode": os.getenv("DOWNLOAD_OU_MODE"),
                "skipPaging": os.getenv("DOWNLOAD_SKIP_PAGING"),
                "format": os.getenv("DOWNLOAD_FORMAT"),
            },
            output_file=os.getenv("PIVOT_INPUT"),
            token=os.getenv("PIVOT_TOKEN"),
        )
    else:
        print("⏩ Téléchargement ignoré")

    # ======================================================
    # ÉTAPE 2 : PIVOT
    # ======================================================
    if do_pivot:
        print("\n📊 ÉTAPE : Pivot & Excel\n")

        run_pivot_and_excel(
            tracked_input=os.getenv("TRACKED_OUTPUT"),
            stage_input=os.getenv("PIVOT_INPUT"),
            output=os.getenv("MERGED_PIVOT_OUTPUT"),
            base_url=os.getenv("PIVOT_BASE_URL"),
            token=os.getenv("PIVOT_TOKEN"),
            program=os.getenv("DOWNLOAD_PROGRAM"),
            aggfunc=os.getenv("PIVOT_AGGFUNC", "first"),
            mapping_file=os.getenv("PIVOT_MAPPING_FILE"),
            state_file=os.getenv("PIVOT_STATE_FILE"),
            strict=False,
        )
    else:
        print("⏩ Pivot ignoré")

    print("\n✅ Pipeline terminé")


if __name__ == "__main__":
    main()
